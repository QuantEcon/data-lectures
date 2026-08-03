"""
Builder for lectures/japan_population_by_age.csv.

Extracts the population of Japan by single year of age from Table 1 of the
Population Estimates published by the Statistics Bureau of Japan, as of
1 October 2024.

The upstream workbook is laid out for printing: ages run down a left-hand block
and continue in a second block to its right, under two levels of merged
headers. This builder flattens that into one row per age.

Counts are published in thousands and are left in those units, since that is
what the source states. The final age, 100, means "100 and over".

Stages: fetch -> pre-process -> validate -> write.
"""

import io
import os
import re
import urllib.request

import openpyxl
import pandas as pd

CURRENT_FILE_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(CURRENT_FILE_DIR)
PUBLISHED_DIR = os.path.join(REPO_ROOT, 'lectures')

SOURCE_URL = 'https://www.stat.go.jp/data/jinsui/2024np/zuhyou/05k2024-1.xlsx'

OUT_FILE = 'japan_population_by_age.csv'

SHEET = '第1表'

# (age label column, first data column) for the left and right blocks of the
# printed table. Data columns run: total both sexes, male, female, sex ratio,
# Japanese both sexes, male, female.
BLOCKS = ((1, 2), (10, 11))
TOTAL_OFFSET = 0        # total population, both sexes
JAPANESE_OFFSET = 4     # Japanese-national population, both sexes

# The "total" row of the source table, used to check that we read the columns
# we think we did.
PUBLISHED_TOTALS = {'total_population': 123802, 'japanese_population': 120296}

# Every figure, including the total, is rounded to the nearest thousand
# independently, so the per-age figures need not sum exactly to the published
# total. 101 roundings of up to half a thousand each bound the gap at ~50;
# the tolerance below sits inside that and far below any column shift.
ROUNDING_TOLERANCE = 60


def fetch():
    with urllib.request.urlopen(SOURCE_URL) as response:
        return openpyxl.load_workbook(io.BytesIO(response.read()))[SHEET]


def pre_process(ws):
    rows = []
    for age_col, data_col in BLOCKS:
        for r in range(1, ws.max_row + 1):
            label = ws.cell(r, age_col).value
            if label is None:
                continue
            # age labels look like "0  歳", "1", "100  歳以上"; the totals row
            # and any notes carry no leading digits
            match = re.match(r'^(\d+)', str(label).replace(' ', '').replace('　', ''))
            if not match:
                continue
            rows.append({
                'age': int(match.group(1)),
                'total_population': ws.cell(r, data_col + TOTAL_OFFSET).value,
                'japanese_population': ws.cell(r, data_col + JAPANESE_OFFSET).value,
            })
    return pd.DataFrame(rows).sort_values('age').reset_index(drop=True)


def validate(df):
    """Refuse to write anything that is not the shape we expect."""
    assert list(df.columns) == ['age', 'total_population', 'japanese_population']
    assert df['age'].tolist() == list(range(101)), 'ages are not 0 to 100'
    assert not df.isnull().values.any()
    # Column check: our per-age figures must sum to the published totals, up to
    # the rounding slack described above. This catches a shifted column, which
    # is the way this parse would fail.
    for column, published in PUBLISHED_TOTALS.items():
        gap = abs(df[column].sum() - published)
        assert gap <= ROUNDING_TOLERANCE, \
            f'{column} sums to {df[column].sum()}, published total is {published}'


def run():
    df = pre_process(fetch())
    validate(df)
    df.to_csv(os.path.join(PUBLISHED_DIR, OUT_FILE), index=False)
    print(f'wrote {OUT_FILE}: {len(df)} rows, '
          f'{df["total_population"].sum():,} thousand people')


if __name__ == '__main__':
    run()
